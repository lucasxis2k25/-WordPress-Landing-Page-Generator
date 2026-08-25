from __future__ import annotations

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
INPUT_JSON = ROOT / "outputs" / "pilot_10_aprovado" / "pilot_results_aprovado.json"
TERMS_JSON = ROOT / "outputs" / "pilot_10_aprovado" / "approved_terms.json"
QUEUE_JSON = ROOT / "outputs" / "pilot_10_aprovado" / "research_queue.json"
OUTPUT_EXCEL = ROOT / "PILOTO_10_PRODUTOS_MAPEAMENTO.xlsx"

HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F497D")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
REGULAR_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def style_header_row(ws, row_idx: int, col_count: int):
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def auto_fit_columns(ws, max_len_cap=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), max_len_cap)

def build_workbook():
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo Executivo"
    
    with open(INPUT_JSON, encoding="utf-8") as f:
        products_data = json.load(f)
    with open(TERMS_JSON, encoding="utf-8") as f:
        terms_data = json.load(f)
    with open(QUEUE_JSON, encoding="utf-8") as f:
        queue_data = json.load(f)

    # Sheet 1: Resumo Executivo
    ws_summary.cell(row=1, column=1, value="Mapeamento ABC — Piloto 10 Produtos").font = TITLE_FONT
    ws_summary.cell(row=2, column=1, value="Resultados validados e auditáveis por produto e família técnica").font = SUBTITLE_FONT
    
    headers_summary = ["Produto", "Família Técnica", "Qtd Total", "Qtd Elegível", "Qtd Comprovada", "Qtd Sob Revisão", "Clientes Total", "Clientes Comprovados", "Clientes Sob Revisão", "% Comprovado"]
    for col_idx, h in enumerate(headers_summary, start=1):
        ws_summary.cell(row=4, column=col_idx, value=h)
    style_header_row(ws_summary, 4, len(headers_summary))
    
    for row_idx, p in enumerate(products_data, start=5):
        s = p["summary"]
        pub_pct = (s["quantity_publicable"] / s["quantity_eligible"]) if s["quantity_eligible"] > 0 else 0.0
        ws_summary.cell(row=row_idx, column=1, value=p["product"]).font = BOLD_FONT
        ws_summary.cell(row=row_idx, column=2, value=p["technical_family"]).font = REGULAR_FONT
        ws_summary.cell(row=row_idx, column=3, value=s["quantity_total"]).font = REGULAR_FONT
        ws_summary.cell(row=row_idx, column=4, value=s["quantity_eligible"]).font = REGULAR_FONT
        ws_summary.cell(row=row_idx, column=5, value=s["quantity_publicable"]).font = REGULAR_FONT
        ws_summary.cell(row=row_idx, column=6, value=s["quantity_under_review"]).font = REGULAR_FONT
        ws_summary.cell(row=row_idx, column=7, value=s["eligible_clients"]).font = REGULAR_FONT
        ws_summary.cell(row=row_idx, column=8, value=s["publicable_clients"]).font = REGULAR_FONT
        ws_summary.cell(row=row_idx, column=9, value=s["review_clients"]).font = REGULAR_FONT
        cell_pct = ws_summary.cell(row=row_idx, column=10, value=pub_pct)
        cell_pct.number_format = "0.0%"
        cell_pct.font = REGULAR_FONT
        for col_idx in range(1, 11):
            ws_summary.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    auto_fit_columns(ws_summary)

    # Sheet 2: Termos Específicos Aprovados
    ws_terms = wb.create_sheet(title="Termos Específicos")
    ws_terms.cell(row=1, column=1, value="Termos Específicos Aprovados (Top 5 por Dimensão)").font = TITLE_FONT
    headers_terms = ["Produto", "Família Técnica", "Dimensão", "Termo Específico", "Status"]
    for col_idx, h in enumerate(headers_terms, start=1):
        ws_terms.cell(row=3, column=col_idx, value=h)
    style_header_row(ws_terms, 3, len(headers_terms))

    for row_idx, t in enumerate(terms_data, start=4):
        ws_terms.cell(row=row_idx, column=1, value=t.get("product", "")).font = REGULAR_FONT
        ws_terms.cell(row=row_idx, column=2, value=t.get("technical_family", "")).font = REGULAR_FONT
        ws_terms.cell(row=row_idx, column=3, value=t.get("dimension", "")).font = REGULAR_FONT
        ws_terms.cell(row=row_idx, column=4, value=t.get("term", "")).font = BOLD_FONT
        ws_terms.cell(row=row_idx, column=5, value=t.get("status", "")).font = REGULAR_FONT
        for col_idx in range(1, 6):
            ws_terms.cell(row=row_idx, column=col_idx).border = THIN_BORDER
    auto_fit_columns(ws_terms)

    # Sheet 3: Clientes Piloto Mapeados
    ws_clients = wb.create_sheet(title="Base Clientes Piloto")
    ws_clients.cell(row=1, column=1, value="Base Completa de Clientes Mapeados — Piloto 10 Produtos").font = TITLE_FONT
    headers_clients = ["Produto", "Família Técnica", "Cliente", "Quantidade", "Curva ABC", "Tipo", "Mercado", "Aplicação Técnica", "Equipamento Físico", "Confiança", "Status"]
    for col_idx, h in enumerate(headers_clients, start=1):
        ws_clients.cell(row=3, column=col_idx, value=h)
    style_header_row(ws_clients, 3, len(headers_clients))

    current_row = 4
    for p in products_data:
        for c in p.get("clients", []):
            ws_clients.cell(row=current_row, column=1, value=p["product"]).font = REGULAR_FONT
            ws_clients.cell(row=current_row, column=2, value=p["technical_family"]).font = REGULAR_FONT
            ws_clients.cell(row=current_row, column=3, value=c.get("client", "")).font = BOLD_FONT
            ws_clients.cell(row=current_row, column=4, value=c.get("quantity", 0)).font = REGULAR_FONT
            ws_clients.cell(row=current_row, column=5, value=c.get("abc", "")).font = BOLD_FONT
            ws_clients.cell(row=current_row, column=6, value=c.get("type", "")).font = REGULAR_FONT
            ws_clients.cell(row=current_row, column=7, value=c.get("market", "")).font = REGULAR_FONT
            ws_clients.cell(row=current_row, column=8, value=c.get("application", "")).font = REGULAR_FONT
            ws_clients.cell(row=current_row, column=9, value=c.get("equipment", "")).font = REGULAR_FONT
            ws_clients.cell(row=current_row, column=10, value=c.get("confidence", "")).font = REGULAR_FONT
            ws_clients.cell(row=current_row, column=11, value=c.get("status", "")).font = REGULAR_FONT
            for col_idx in range(1, 12):
                ws_clients.cell(row=current_row, column=col_idx).border = THIN_BORDER
            current_row += 1
    auto_fit_columns(ws_clients)

    # Sheet 4: Fila de Revisão (Research Queue)
    ws_queue = wb.create_sheet(title="Fila de Pesquisa")
    ws_queue.cell(row=1, column=1, value="Fila de Validação de Evidências (Clientes sem Cache)").font = TITLE_FONT
    headers_queue = ["Chave Única", "Cliente", "Família Técnica", "Exemplo de Produto", "Prioridade ABC", "Motivo", "Status"]
    for col_idx, h in enumerate(headers_queue, start=1):
        ws_queue.cell(row=3, column=col_idx, value=h)
    style_header_row(ws_queue, 3, len(headers_queue))

    for row_idx, q in enumerate(queue_data, start=4):
        ws_queue.cell(row=row_idx, column=1, value=q.get("unique_key", "")).font = REGULAR_FONT
        ws_queue.cell(row=row_idx, column=2, value=q.get("client", "")).font = BOLD_FONT
        ws_queue.cell(row=row_idx, column=3, value=q.get("technical_family", "")).font = REGULAR_FONT
        ws_queue.cell(row=row_idx, column=4, value=q.get("product_example", "")).font = REGULAR_FONT
        ws_queue.cell(row=row_idx, column=5, value=q.get("priority", "")).font = BOLD_FONT
        ws_queue.cell(row=row_idx, column=6, value=q.get("reason", "")).font = REGULAR_FONT
        ws_queue.cell(row=row_idx, column=7, value=q.get("status", "")).font = REGULAR_FONT
        for col_idx in range(1, 8):
            ws_queue.cell(row=row_idx, column=col_idx).border = THIN_BORDER
    auto_fit_columns(ws_queue)

    wb.save(OUTPUT_EXCEL)
    print(f"Excel gerado com sucesso em: {OUTPUT_EXCEL}")

if __name__ == "__main__":
    build_workbook()
