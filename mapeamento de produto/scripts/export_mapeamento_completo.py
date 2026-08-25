"""Gera Excel com mapeamento individual por produto, no formato da aba VENT FS4-400 ET.

Cada produto recebe sua própria aba com:
- Resumo do produto
- Curva ABC de Clientes
- Curva ABC de Mercados com 5 termos principais
- Curva ABC de Aplicações com 5 termos técnicos
- Curva ABC de Equipamentos com 5 termos principais
"""
from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from mapping.abc import abc_class, abc_table, top_terms
from mapping.parquet_io import read_parquet
from mapping.rules import eligible_client, normalize_text, client_type

# --- Cores e estilos ---
DARK_TEAL_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
MEDIUM_TEAL_FILL = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="1A5276")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="5D6D7E")
BOLD_10 = Font(name="Calibri", size=10, bold=True)
REGULAR_10 = Font(name="Calibri", size=10)
REGULAR_9 = Font(name="Calibri", size=9)
REVIEW_FONT = Font(name="Calibri", size=10, italic=True, color="C0392B")

THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# --- Produtos a processar (excluindo Tangencial, Plus Unic, Conecto, ECM sem dados) ---
PRODUCTS = [
    "VENT. FF/2-146 P 220V",
    "VENT. FS/4-400 EMBT - 230 V",
    "MICROVENTILADOR A17251VBHBL - 110/220V",
    "VENT.FB/2-190MCD- 230V. 50/60",
    "VENT. FS/2-300 EM -  230 V",
    "VENT. FF/2-160 VS- 220V. 60 HZ",
    "VENT. ECM 30 - 50/60HZ",
]

# Nomes curtos para abas Excel (max 31 chars)
SHEET_NAMES = {
    "VENT. FF/2-146 P 220V": "Map FF2-146P",
    "VENT. FS/4-400 EMBT - 230 V": "Map FS4-400 EMBT",
    "MICROVENTILADOR A17251VBHBL - 110/220V": "Map MICRO A17251",
    "VENT.FB/2-190MCD- 230V. 50/60": "Map FB2-190MCD",
    "VENT. FS/2-300 EM -  230 V": "Map FS2-300 EM",
    "VENT. FF/2-160 VS- 220V. 60 HZ": "Map FF2-160VS",
    "VENT. ECM 30 - 50/60HZ": "Map ECM 30",
}


def _style_header(ws, row: int, cols: int, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill or DARK_TEAL_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _write_row(ws, row: int, values: list, fonts: list | None = None, aligns: list | None = None, fill=None):
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = (fonts[c - 1] if fonts and c - 1 < len(fonts) else REGULAR_10)
        cell.alignment = (aligns[c - 1] if aligns and c - 1 < len(aligns) else ALIGN_LEFT)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _auto_width(ws, max_w=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        mx = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[letter].width = min(max(mx + 3, 12), max_w)


def _aggregate_terms_for_group(rows_in_group: list[dict], dimension: str, cache: dict, product_family: str = "") -> str:
    """Agrega os 5 termos granulares mais frequentes de um grupo ABC.
    
    Aplica portão anti-canibalização: termos de refrigeração são filtrados
    para produtos que NÃO são AXIAL (refrigeração).
    """
    all_terms: list[str] = []
    is_refrig_product = product_family == "AXIAL"

    refrig_term_markers = [
        "evaporador", "frigor", "condensador", "plug-in", "congelamento",
        "forcador", "forçador", "câmara fria", "cadeia do frio",
        "refrigera", "túnel de resfriamento", "tunnel", "baú", "bau",
        "unidade condensadora", "monobloco", "chillers", "chiller",
        "torre de resfriamento", "dry cooler", "expositor refrigerado",
        "freezer", "balcão refrigerado", "bebedouro", "cervejeira",
        "refresqueira", "câmara frigor", "unid de refrig",
    ]

    def _term_is_contaminated(term: str) -> bool:
        if is_refrig_product:
            return False
        t_lower = term.lower()
        return any(marker in t_lower for marker in refrig_term_markers)

    for r in rows_in_group:
        key = normalize_text(r.get("client_raw"))
        cached = cache.get(key)

        if dimension == "market":
            if cached and cached.get("terms_market"):
                for t in str(cached["terms_market"]).split("; "):
                    t = t.strip()
                    if t and not t.isupper() and len(t) > 3 and not _term_is_contaminated(t):
                        all_terms.append(t)
            seg = str(r.get("segment") or "").strip()
            if seg and seg not in ("", "OUTRAS APLICACOES"):
                term = seg.title().replace("E ", "e ").replace("De ", "de ")
                if not _term_is_contaminated(term):
                    all_terms.append(term)

        elif dimension == "application":
            if cached and cached.get("terms_application"):
                for t in str(cached["terms_application"]).split("; "):
                    t = t.strip()
                    if t and len(t) > 3 and not _term_is_contaminated(t):
                        all_terms.append(t)
            for col in ["equipment_candidate_1", "equipment_candidate_2"]:
                val = str(r.get(col) or "").strip()
                if val and "provável" not in val.lower() and not _term_is_contaminated(val):
                    all_terms.append(val)

        elif dimension == "equipment":
            if cached and cached.get("terms_equipment"):
                for t in str(cached["terms_equipment"]).split("; "):
                    t = t.strip()
                    if t and len(t) > 3 and not t.startswith("http") and t not in ("Alta", "Média", "Baixa"):
                        if not _term_is_contaminated(t):
                            all_terms.append(t)
            for col in ["equipment_candidate_1", "equipment_candidate_2", "equipment_candidate_3", "equipment_candidate_4", "equipment_candidate_5"]:
                val = str(r.get(col) or "").strip()
                if val and "provável" not in val.lower() and not _term_is_contaminated(val):
                    all_terms.append(val)

    filtered: list[str] = []
    generic_prefixes = ("Revisar", "Não comprovado", "Validado", "Base interna", "Site oficial", "Cadastro")
    for t in all_terms:
        if any(t.startswith(gp) for gp in generic_prefixes):
            continue
        if t.startswith("http"):
            continue
        filtered.append(t)

    counts = Counter(filtered)
    best = [term for term, _ in counts.most_common(5)]
    return "; ".join(best)


# --- Portao Anti-Canibalizacao ---
_REFRIG_MARKERS = [
    "evaporador", "frigor", "condensador", "plug-in",
    "refrigeracao", "cadeia do frio", "forcador",
    "congelamento", "tunnel", "bau",
]
_HEAVY_EQUIP_MARKERS = [
    "evaporador de camara", "evaporador de câmara",
    "túnel de", "tunnel de", "unidade condensadora",
    "baú", "bau", "carroceria", "compressor",
]


def _cache_is_contaminated(cached: dict, product_family: str) -> bool:
    """True se cache veio de familia de refrigeracao e produto atual nao e AXIAL."""
    origin = str(cached.get("family_id") or "")
    is_refrig_cache = origin in ("A12038", "VENT_FS4_400_ET")
    is_refrig_product = product_family == "AXIAL"
    if not is_refrig_cache or is_refrig_product:
        return False
    for field in ("market", "application", "equipment"):
        val = str(cached.get(field) or "").lower()
        for marker in _REFRIG_MARKERS:
            if marker in val:
                return True
    return False


def _micro_equip_contaminated(equipment: str) -> bool:
    """Microventiladores nao atuam em equipamento pesado de refrigeracao."""
    eq_lower = equipment.lower()
    for marker in _HEAVY_EQUIP_MARKERS:
        if marker in eq_lower:
            return True
    return False


def _safe_cache_or_parquet(cached: dict | None, row: dict, product_family: str) -> tuple[str, str, str, str, str]:
    """Retorna (market, application, equipment, confidence, status) usando cache apenas se nao contaminado."""
    if cached and not _cache_is_contaminated(cached, product_family):
        market = cached.get("market") or row.get("market_canonical") or row.get("market_base") or "Revisar"
        application = cached.get("application") or "Revisar"
        equipment = cached.get("equipment") or row.get("equipment_candidate_1") or "Revisar"
        if "MICRO" in product_family and _micro_equip_contaminated(equipment):
            equipment = row.get("equipment_candidate_1") or "Revisar"
            application = "Revisar"
        confidence = str(cached.get("confidence", "Baixa"))
        status = "Comprovado"
    else:
        market = row.get("market_canonical") or row.get("market_base") or "Revisar"
        application = "Revisar"
        equipment = row.get("equipment_candidate_1") or "Revisar"
        confidence = "Baixa"
        status = "Revisar" if not cached else "Revisar (cache bloqueado)"
    return market, application, equipment, confidence, status


def _build_product_abc(eligible_rows: list[dict], cache: dict, product_family: str = "") -> dict[str, list[dict]]:
    """Constrói as curvas ABC com 5 termos para um produto."""
    # Agrupar linhas por chave para cada dimensão
    def _group_rows(rows, key_fn):
        groups: dict[str, list[dict]] = defaultdict(list)
        totals: dict[str, float] = defaultdict(float)
        for r in rows:
            label = key_fn(r)
            qty = float(r.get("quantity") or 0)
            groups[label].append(r)
            totals[label] += qty
        return groups, totals

    def _market_key(r):
        key = normalize_text(r.get("client_raw"))
        c = cache.get(key)
        market, _, _, _, _ = _safe_cache_or_parquet(c, r, product_family)
        return market

    def _application_key(r):
        key = normalize_text(r.get("client_raw"))
        c = cache.get(key)
        _, application, _, _, _ = _safe_cache_or_parquet(c, r, product_family)
        return application

    def _equipment_key(r):
        key = normalize_text(r.get("client_raw"))
        c = cache.get(key)
        _, _, equipment, _, _ = _safe_cache_or_parquet(c, r, product_family)
        return equipment

    result = {}
    for dimension, key_fn in [("markets", _market_key), ("applications", _application_key), ("equipment", _equipment_key)]:
        groups, totals = _group_rows(eligible_rows, key_fn)
        grand_total = sum(totals.values())
        sorted_labels = sorted(totals.keys(), key=lambda l: (-totals[l], l))
        abc_rows = []
        accumulated = 0.0
        for rank, label in enumerate(sorted_labels, start=1):
            qty = totals[label]
            share = qty / grand_total if grand_total else 0.0
            accumulated += share
            dim_key = dimension.rstrip("s") if dimension != "equipment" else "equipment"
            terms = _aggregate_terms_for_group(groups[label], dim_key, cache, product_family=product_family)
            abc_rows.append({
                "rank": rank,
                "label": label,
                "quantity": qty,
                "share": share,
                "accumulated": accumulated,
                "abc": abc_class(accumulated),
                "terms": terms,
            })
        result[dimension] = abc_rows
    return result


def _build_client_abc(eligible_rows: list[dict], cache: dict, product_family: str = "") -> list[dict]:
    """Constrói a curva ABC de clientes com mercado/aplicação/equipamento individuais."""
    by_client: dict[str, dict] = {}
    for r in eligible_rows:
        raw = str(r.get("client_raw") or "")
        key = normalize_text(raw)
        qty = float(r.get("quantity") or 0)
        if key not in by_client:
            c = cache.get(key)
            market, application, equipment, confidence, status = _safe_cache_or_parquet(c, r, product_family)
            by_client[key] = {
                "client": raw,
                "quantity": 0.0,
                "type": (c.get("client_type") if c else None) or client_type(r.get("channel")),
                "market": market,
                "application": application,
                "equipment": equipment,
                "confidence": confidence,
                "status": status,
            }
        by_client[key]["quantity"] += qty

    sorted_clients = sorted(by_client.values(), key=lambda x: (-x["quantity"], x["client"]))
    grand_total = sum(c["quantity"] for c in sorted_clients)
    accumulated = 0.0
    result = []
    for rank, c in enumerate(sorted_clients, start=1):
        share = c["quantity"] / grand_total if grand_total else 0.0
        accumulated += share
        result.append({
            **c,
            "rank": rank,
            "share": share,
            "accumulated": accumulated,
            "abc": abc_class(accumulated),
        })
    return result


def write_product_sheet(wb, product: str, all_rows: list[dict], cache: dict):
    """Escreve a aba completa de um produto no workbook."""
    sheet_name = SHEET_NAMES.get(product, product[:31])
    ws = wb.create_sheet(title=sheet_name)

    prod_rows = [r for r in all_rows if r.get("product") == product and float(r.get("quantity") or 0) > 0]
    eligible = [r for r in prod_rows if eligible_client(r.get("channel"), r.get("profile"), r.get("segment"))]
    excluded = [r for r in prod_rows if r not in eligible]
    family = str(eligible[0].get("technical_family", "")) if eligible else str(prod_rows[0].get("technical_family", "")) if prod_rows else ""

    qty_total = sum(float(r.get("quantity") or 0) for r in prod_rows)
    qty_eligible = sum(float(r.get("quantity") or 0) for r in eligible)
    qty_excluded = sum(float(r.get("quantity") or 0) for r in excluded)

    client_abc = _build_client_abc(eligible, cache, product_family=family)
    dimension_abc = _build_product_abc(eligible, cache, product_family=family)

    n_clients_a = sum(1 for c in client_abc if c["abc"] == "A")
    n_clients_b = sum(1 for c in client_abc if c["abc"] == "B")
    n_clients_c = sum(1 for c in client_abc if c["abc"] == "C")
    review_count = sum(1 for c in client_abc if c["status"] == "Revisar")

    # --- Cabeçalho do produto ---
    row = 1
    ws.cell(row=row, column=1, value=f"Mapeamento ABC — {product}").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Família Técnica: {family}").font = SUBTITLE_FONT
    row += 2

    # --- Resumo numérico ---
    ws.cell(row=row, column=1, value="Resumo do Produto").font = BOLD_10
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    summary_data = [
        ("Quantidade Total (prefixo)", qty_total),
        ("Quantidade Elegível", qty_eligible),
        ("Quantidade Excluída (Revenda/Manutenção)", qty_excluded),
        ("Total de Clientes Elegíveis", len(client_abc)),
        ("Clientes A", n_clients_a),
        ("Clientes B", n_clients_b),
        ("Clientes C", n_clients_c),
        ("Clientes Sob Revisão", review_count),
    ]
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = REGULAR_10
        ws.cell(row=row, column=2, value=value).font = BOLD_10
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # --- Curva ABC de Clientes ---
    row += 1
    ws.cell(row=row, column=1, value="Curva ABC — Clientes").font = BOLD_10
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    client_headers = ["Rank", "Cliente", "Quantidade", "Participação %", "Acumulado %", "ABC", "Mercado", "Aplicação Técnica", "Equipamento Físico"]
    for c, h in enumerate(client_headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, len(client_headers), fill=MEDIUM_TEAL_FILL)
    row += 1

    for c in client_abc:
        is_review = c["status"] == "Revisar"
        fill = REVIEW_FILL if is_review else None
        font_name = REVIEW_FONT if is_review else REGULAR_10
        vals = [
            c["rank"], c["client"], c["quantity"],
            f"{c['share']:.1%}", f"{c['accumulated']:.1%}", c["abc"],
            c["market"], c["application"], c["equipment"]
        ]
        fonts = [REGULAR_10, BOLD_10, REGULAR_10, REGULAR_10, REGULAR_10, BOLD_10, font_name, font_name, font_name]
        _write_row(ws, row, vals, fonts=fonts, fill=fill)
        row += 1

    # --- Curvas ABC de Dimensões (Mercados, Aplicações, Equipamentos) ---
    dim_configs = [
        ("Curva ABC — Mercados", "markets", "Mercado", "5 termos principais"),
        ("Curva ABC — Aplicações", "applications", "Aplicação", "5 termos técnicos"),
        ("Curva ABC — Produtos / Equipamentos", "equipment", "Produto / equipamento", "5 termos principais"),
    ]

    for title, dim_key, label_header, terms_header in dim_configs:
        row += 2
        ws.cell(row=row, column=1, value=title).font = BOLD_10
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        dim_headers = ["Rank", label_header, "Quantidade", "Participação %", "Acumulado %", "ABC", terms_header]
        for c, h in enumerate(dim_headers, start=1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(dim_headers))
        row += 1

        for item in dimension_abc[dim_key]:
            is_review = "Revisar" in item["label"]
            fill = REVIEW_FILL if is_review else (LIGHT_GRAY_FILL if item["rank"] % 2 == 0 else None)
            vals = [
                item["rank"], item["label"], item["quantity"],
                f"{item['share']:.1%}", f"{item['accumulated']:.1%}", item["abc"],
                item["terms"]
            ]
            fonts = [REGULAR_10, BOLD_10, REGULAR_10, REGULAR_10, REGULAR_10, BOLD_10, REGULAR_9]
            _write_row(ws, row, vals, fonts=fonts, fill=fill)
            row += 1

    _auto_width(ws)
    # Ajustar colunas de termos para ser mais larga
    ws.column_dimensions["G"].width = 65
    ws.column_dimensions["H"].width = 50
    ws.column_dimensions["I"].width = 50


def main():
    print("Carregando dados...")
    all_rows = read_parquet(ROOT / "data" / "normalized" / "consolidacao.parquet")
    master_rows = read_parquet(ROOT / "data" / "master" / "client_master.parquet")
    cache: dict[str, dict] = {}
    for r in master_rows:
        cache.setdefault(str(r.get("client_key") or ""), r)

    # Filtrar apenas produtos que possuem dados elegíveis
    active_products = []
    for p in PRODUCTS:
        prod_rows = [r for r in all_rows if r.get("product") == p and float(r.get("quantity") or 0) > 0]
        eligible = [r for r in prod_rows if eligible_client(r.get("channel"), r.get("profile"), r.get("segment"))]
        if eligible:
            active_products.append(p)
            print(f"  {p}: {len(eligible)} clientes elegíveis")
        else:
            print(f"  {p}: SEM DADOS ELEGÍVEIS — ignorando")

    wb = openpyxl.Workbook()

    # Aba Resumo
    ws_resume = wb.active
    ws_resume.title = "Resumo Geral"
    ws_resume.cell(row=1, column=1, value="Mapeamento ABC — Piloto de Produtos").font = TITLE_FONT
    ws_resume.cell(row=2, column=1, value="Cada produto possui sua própria aba com curvas ABC e termos específicos").font = SUBTITLE_FONT

    resume_headers = ["Produto", "Família Técnica", "Qtd Total", "Qtd Elegível", "Clientes Elegíveis", "Status"]
    row = 4
    for c, h in enumerate(resume_headers, start=1):
        ws_resume.cell(row=row, column=c, value=h)
    _style_header(ws_resume, row, len(resume_headers))
    row += 1

    for p in active_products:
        prod_rows = [r for r in all_rows if r.get("product") == p and float(r.get("quantity") or 0) > 0]
        eligible = [r for r in prod_rows if eligible_client(r.get("channel"), r.get("profile"), r.get("segment"))]
        family = eligible[0].get("technical_family", "") if eligible else ""
        qty_total = sum(float(r.get("quantity") or 0) for r in prod_rows)
        qty_eligible = sum(float(r.get("quantity") or 0) for r in eligible)
        n_clients = len({normalize_text(r.get("client_raw")) for r in eligible})
        _write_row(ws_resume, row, [p, family, qty_total, qty_eligible, n_clients, f"Ver aba '{SHEET_NAMES.get(p, p[:31])}'"],
                    fonts=[BOLD_10, REGULAR_10, REGULAR_10, REGULAR_10, REGULAR_10, SUBTITLE_FONT])
        row += 1
    _auto_width(ws_resume)

    # Abas individuais
    for p in active_products:
        print(f"Gerando aba: {SHEET_NAMES.get(p, p[:31])} ...")
        write_product_sheet(wb, p, all_rows, cache)

    output = ROOT / "MAPEAMENTO_PILOTO_COMPLETO.xlsx"
    wb.save(output)
    print(f"\nExcel gerado com sucesso: {output}")


if __name__ == "__main__":
    main()
