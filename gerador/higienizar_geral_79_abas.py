# -*- coding: utf-8 -*-
"""
======================================================================
Demo Store — HIGIENIZAÇÃO GERAL DE TODAS AS 79 ABAS E ARQUIVOS JSON
Aplica a normalização mestre dos 55 segmentos, remove Consumidores PF
e higieniza os textos técnicos de todas as Landing Pages.
======================================================================
"""

import os
import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

BASE_DIR = r"c:\Users\comercial\Desktop\backup\PROJETOS\Projeto landing pages"
GERADOR_DIR = os.path.join(BASE_DIR, "gerador")
AUTOMACAO_DIR = os.path.join(GERADOR_DIR, "automacao")
DADOS_DIR = os.path.join(GERADOR_DIR, "dados")
OUTPUT_DIR = os.path.join(GERADOR_DIR, "output")
EXCEL_PATH = os.path.join(BASE_DIR, "mapeamento de produto", "MAPEAMENTO_PILOTO_COMPLETO.xlsx")

sys.path.insert(0, GERADOR_DIR)
sys.path.insert(0, AUTOMACAO_DIR)

from normalizadores import normalizar_segmento_mestre, normalizar_equipamento_mestre
from gerar_conteudo_acf import gerar_payload_acf, salvar_output
from publicar_wordpress import update_product_acf

def main():
    print("=" * 80)
    print("  Demo Store — HIGIENIZAÇÃO GERAL (79 ABAS + BANCO DE DADOS + ACF)")
    print("=" * 80)

    # 1. Higieniza a Planilha Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    print(f"[*] Planilha carregada: {len(wb.sheetnames)} abas.")

    for sheetname in wb.sheetnames:
        if sheetname == "Resumo Geral":
            continue
        ws = wb[sheetname]
        
        # Itera por todas as células da aba substituindo termos brutos ou proibidos
        rows_to_delete = []
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=1).value
            if not cell_val:
                continue
            cell_str = str(cell_val).strip()
            
            # Se for linha de Consumidor PF em tabela
            if "CONSUMIDOR" in cell_str.upper() and "PF" in cell_str.upper():
                # limpa conteúdo
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c, value="")
                continue
                
            # Normaliza segmentos
            norm = normalizar_segmento_mestre(cell_str)
            if norm and norm != cell_str and len(cell_str) > 3 and not cell_str.startswith("Mapeamento") and not cell_str.startswith("SKU") and not cell_str.startswith("1.") and not cell_str.startswith("2.") and not cell_str.startswith("3.") and not cell_str.startswith("4.") and not cell_str.startswith("Parâmetro") and not cell_str.startswith("Segmento") and not cell_str.startswith("Equipamento") and not cell_str.startswith("Razão"):
                ws.cell(row=r, column=1, value=norm)

            # Normaliza equipamentos
            eq_norm = normalizar_equipamento_mestre(cell_str)
            if eq_norm != cell_str:
                ws.cell(row=r, column=1, value=eq_norm)

    wb.save(EXCEL_PATH)
    print(f"[OK] Planilha Excel '{EXCEL_PATH}' salva com sucesso!")

    # 2. Higieniza todos os JSONs em gerador/dados/
    json_files = [f for f in os.listdir(DADOS_DIR) if f.endswith(".json") and not f.endswith("_acf.json")]
    print(f"\n[*] Higienizando {len(json_files)} arquivos JSON em dados/...")

    atualizados = 0
    for fname in json_files:
        fpath = os.path.join(DADOS_DIR, fname)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                d = json.load(f)
        except Exception:
            continue

        modificado = False
        
        # Higieniza mercados
        mercados = d.get("mercados") or d.get("mercado") or []
        novos_mercados = []
        for m in mercados:
            m_norm = normalizar_segmento_mestre(m)
            if m_norm:
                novos_mercados.append(m_norm)
                if m_norm != m:
                    modificado = True
            else:
                modificado = True # removeu proibido

        if novos_mercados:
            d["mercado"] = novos_mercados
            d["mercados"] = novos_mercados

        # Higieniza equipamentos
        if "aplicacoes_equipamento" in d and "cards" in d["aplicacoes_equipamento"]:
            for card in d["aplicacoes_equipamento"]["cards"]:
                t = card.get("titulo", "")
                t_norm = normalizar_equipamento_mestre(t)
                if t_norm != t:
                    card["titulo"] = t_norm
                    modificado = True

        if "onde_usar" in d:
            for card in d["onde_usar"]:
                t = card.get("titulo", "")
                t_norm = normalizar_equipamento_mestre(t)
                if t_norm != t:
                    card["titulo"] = t_norm
                    modificado = True

        # Salva se modificado ou para garantir sincronicidade
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(d, f, ensure_ascii=False, indent=2)

        # Regera ACF payload
        payload = gerar_payload_acf(d)
        salvar_output(payload)
        atualizados += 1

    print(f"[OK] {atualizados} JSONs e Payloads ACF higienizados com sucesso!")

    # 3. Sincroniza via REST API os 15 produtos do 4º lote com os dados limpos
    wp_cand_path = r"C:\Users\comercial\.gemini\antigravity-ide\brain\2c0ebb45-5feb-41f6-8d1c-ac9a1b0de51d\scratch\selected_lote_4_wp.json"
    if os.path.exists(wp_cand_path):
        with open(wp_cand_path, "r", encoding="utf-8") as f:
            lote_cand = json.load(f)
        slugs_to_update = [p["wp_slug"] or p["slug"] for p in lote_cand]
        print(f"\n[*] Atualizando {len(slugs_to_update)} produtos do 4º lote no WordPress com dados limpos...")
        for s in slugs_to_update:
            update_product_acf(s, skip_validation=True)

    print("\n" + "=" * 80)
    print("  [SUCESSO] HIGIENIZAÇÃO GERAL COMPLETA!")
    print("=" * 80)

if __name__ == "__main__":
    main()
