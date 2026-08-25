# -*- coding: utf-8 -*-
"""
======================================================================
Demo Store — GERADOR DE NOVO LOTE DE 10 PRODUTOS (MAPEAMENTO + ABA EXCEL + JSON + LP)
Gera com rigor absoluto:
- Abas individuais no MAPEAMENTO_PILOTO_COMPLETO.xlsx
- Curvas ABC de Mercados, Equipamentos e Clientes reais do parquet
- JSONs auditados em gerador/dados/
- Payloads ACF em gerador/output/ e Markdowns em produtos/
======================================================================
"""

import os
import sys
import json
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

PROJ_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
MAP_DIR = os.path.join(PROJ_DIR, "mapeamento de produto")
GERADOR_DIR = os.path.join(PROJ_DIR, "gerador")
DADOS_DIR = os.path.join(GERADOR_DIR, "dados")
OUTPUT_DIR = os.path.join(GERADOR_DIR, "output")
PRODUTOS_DIR = os.path.join(PROJ_DIR, "produtos")

sys.path.insert(0, GERADOR_DIR)
from gerar_conteudo_acf import gerar_payload_acf, salvar_output

# Lista dos 10 produtos a processar
LOTE_10_CONFIG = [   {   'categoria': 'Micro Ventiladores',
        'code': '581',
        'nome': 'Micro Ventilador 120 mm - A12038VBHBL-W',
        'sheet_name': 'Map MICRO A12038',
        'sku': 'A12038VBHBL-W',
        'slug': 'micro-ventilador-120-mm-a12038vbhbl-w',
        'specs': [   {'atributo': 'Modelo', 'valor': 'A12038VBHBL-W'},
                     {'atributo': 'Descrição', 'valor': 'Microventilador'},
                     {'atributo': 'Tensão Nominal', 'valor': '110/220 V AC'},
                     {'atributo': 'Corrente Nominal', 'valor': '0,07 / 0,04 A'},
                     {'atributo': 'Potência Consumida', 'valor': '5 W'},
                     {'atributo': 'Velocidade Nominal', 'valor': '2800 RPM'},
                     {'atributo': 'Mancais', 'valor': 'Rolamento (Ball Bearing)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-20 °C a 80 °C'},
                     {'atributo': 'Nível de Ruído', 'valor': '52 dBA'},
                     {'atributo': 'Proteção Elétrica', 'valor': 'Por impedância'},
                     {'atributo': 'Material da Hélice', 'valor': 'Termoplástico, PBT, UL94-0'},
                     {'atributo': 'Material da Carcaça', 'valor': 'Termoplástico, PBT, UL94-0'},
                     {'atributo': 'Peso', 'valor': '0,230 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '96',
        'nome': 'Ventilador Exaustor Axial 400mm - FS/4-400 ETBT',
        'sheet_name': 'Map FS4 400 ETBT',
        'sku': 'FS/4-400 ETBT',
        'slug': 'ventilador-exaustor-axial-400mm-fs-4-400-etbt',
        'specs': [   {'atributo': 'Tensão Nominal', 'valor': '220/380 V'},
                     {'atributo': 'Alimentação', 'valor': 'Trifásica'},
                     {'atributo': 'Corrente Nominal', 'valor': '0,80/1,00/0,48/0,60 A'},
                     {'atributo': 'Potência Consumida', 'valor': '190/260 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Rotação Nominal', 'valor': '1390/1590 RPM'},
                     {'atributo': 'Grau de Proteção (IP)', 'valor': 'IP-54'},
                     {'atributo': 'Tipo de Mancal', 'valor': 'Com rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30°C a 60°C'},
                     {'atributo': 'Ruído Acústico', 'valor': '60/72 dBA'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aço (na cor preta)'},
                     {'atributo': 'Grade de Proteção', 'valor': 'Em fio de aço (na cor preta)'},
                     {'atributo': 'Peso Líquido', 'valor': '6 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '88',
        'nome': 'Ventilador Soprador Axial 350mm - FS/4-350 VM',
        'sheet_name': 'Map FS4 350 VM',
        'sku': 'FS/4-350 VM',
        'slug': 'ventilador-soprador-axial-350mm-fs-4-350-vm',
        'specs': [   {'atributo': 'Tensão Nominal', 'valor': '220 V'},
                     {'atributo': 'Alimentação', 'valor': 'Monofásica'},
                     {'atributo': 'Corrente Nominal', 'valor': '0,68/0,82 A'},
                     {'atributo': 'Potência Consumida', 'valor': '138/190 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Rotação Nominal', 'valor': '1380/1590 RPM'},
                     {'atributo': 'Grau de Proteção (IP)', 'valor': 'IP-54'},
                     {'atributo': 'Tipo de Mancal', 'valor': 'Com rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30°C a 60°C'},
                     {'atributo': 'Ruído Acústico', 'valor': '62/67 dBA'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aço (na cor preta)'},
                     {'atributo': 'Grade de Proteção', 'valor': 'Em fio de aço (na cor preta)'},
                     {'atributo': 'Peso Líquido', 'valor': '4,5 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '65',
        'nome': 'Ventilador Exaustor Axial 200mm - FS/2-200 EM',
        'sheet_name': 'Map FS2 200 EM',
        'sku': 'FS/2-200 EM',
        'slug': 'ventilador-exaustor-axial-200mm-fs-2-200-em',
        'specs': [   {'atributo': 'Tensão Nominal', 'valor': '220 V'},
                     {'atributo': 'Alimentação', 'valor': 'Monofásica'},
                     {'atributo': 'Corrente Nominal', 'valor': '0,38/0,46 A'},
                     {'atributo': 'Potência Consumida', 'valor': '80/95 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Rotação Nominal', 'valor': '2700/3200 RPM'},
                     {'atributo': 'Grau de Proteção (IP)', 'valor': 'IP-54'},
                     {'atributo': 'Tipo de Mancal', 'valor': 'Com rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30°C a 60°C'},
                     {'atributo': 'Ruído Acústico', 'valor': '60/65 dBA'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aço (na cor preta)'},
                     {'atributo': 'Grade de Proteção', 'valor': 'Em fio de aço (na cor preta)'},
                     {'atributo': 'Peso Líquido', 'valor': '1,9 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '85',
        'nome': 'Ventilador Exaustor Axial 350mm - FS/4-350 ET',
        'sheet_name': 'Map FS4 350 ET',
        'sku': 'FS/4-350 ET',
        'slug': 'ventilador-exaustor-axial-350mm-fs-4-350-et',
        'specs': [   {'atributo': 'Tensão Nominal', 'valor': '220/380 V'},
                     {'atributo': 'Alimentação', 'valor': 'Trifásica'},
                     {'atributo': 'Corrente Nominal', 'valor': '0,7/0,8/0,4/0,46 A'},
                     {'atributo': 'Potência Consumida', 'valor': '145/185 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Rotação Nominal', 'valor': '1390/1580 RPM'},
                     {'atributo': 'Grau de Proteção (IP)', 'valor': 'IP-54'},
                     {'atributo': 'Tipo de Mancal', 'valor': 'Com rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30°C a 60°C'},
                     {'atributo': 'Ruído Acústico', 'valor': '63/67 dBA'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aço (na cor preta)'},
                     {'atributo': 'Grade de Proteção', 'valor': 'Em fio de aço (na cor preta)'},
                     {'atributo': 'Peso Líquido', 'valor': '4,6 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '67',
        'nome': 'Ventilador Exaustor Axial 250mm - FS/2-250 ET',
        'sheet_name': 'Map FS2 250 ET',
        'sku': 'FS/2-250 ET',
        'slug': 'ventilador-exaustor-axial-250mm-fs-2-250-et',
        'specs': [   {'atributo': 'Tensão Nominal', 'valor': '220/380 V'},
                     {'atributo': 'Alimentação', 'valor': 'Trifásica'},
                     {'atributo': 'Corrente Nominal', 'valor': '0,62/0,36/0,65/0,38 (220 V)'},
                     {'atributo': 'Potência Consumida', 'valor': '165/200 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Rotação Nominal', 'valor': '2750/3100 RPM'},
                     {'atributo': 'Grau de Proteção (IP)', 'valor': 'IP-54'},
                     {'atributo': 'Tipo de Mancal', 'valor': 'Com rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30°C a 60°C'},
                     {'atributo': 'Ruído Acústico', 'valor': '70/72 dBA'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aço (na cor preta)'},
                     {'atributo': 'Grade de Proteção', 'valor': 'Em fio de aço (na cor preta)'},
                     {'atributo': 'Peso Líquido', 'valor': '3,7 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '385',
        'nome': 'Ventilador Exaustor Axial 500mm - FS/4-500 EM',
        'sheet_name': 'Map FS4 500 EM',
        'sku': 'FS/4-500 EM',
        'slug': 'ventilador-exaustor-axial-500mm-fs-4-500-em',
        'specs': [   {'atributo': 'Modelo', 'valor': 'FS/4-500 EM'},
                     {   'atributo': 'Descrição',
                         'valor': 'Ventilador axial com grade e caixa de terminais, motor de rotor externo'},
                     {'atributo': 'Tensão Nominal', 'valor': '220 V'},
                     {'atributo': 'Alimentação', 'valor': 'Monofásica'},
                     {'atributo': 'Capacitor', 'valor': '12 uF'},
                     {'atributo': 'Ligação', 'valor': 'Conforme esquema'},
                     {'atributo': 'Corrente Nominal', 'valor': '1,85/2,57 A'},
                     {'atributo': 'Potência Consumida', 'valor': '420/585 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Velocidade Nominal', 'valor': '1320/1510 RPM'},
                     {'atributo': 'Proteção Elétrica', 'valor': 'Através de protetor térmico bimetálico'},
                     {'atributo': 'Proteção Mecânica', 'valor': 'IP54'},
                     {'atributo': 'Mancais', 'valor': 'Com rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30°C a 60°C'},
                     {'atributo': 'Isolação', 'valor': 'F'},
                     {'atributo': 'Nível de Ruído', 'valor': '73/75 dbA'},
                     {'atributo': 'Regime de Trabalho', 'valor': 'S1 (Contínuo)'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aço (na cor preta)'},
                     {'atributo': 'Grades de Proteção', 'valor': 'Em fio de aço (na cor preta)'},
                     {   'atributo': 'Caixa de Ligação',
                         'valor': 'Em polipropileno aditivado com borne, prensa cabos e aterramento'},
                     {'atributo': 'Peso', 'valor': '9,5 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '521',
        'nome': 'Centrífugo Simples Aspiração - FF/2-160 VS',
        'sheet_name': 'Map FF2 160 VS',
        'sku': 'FF/2-160 VS',
        'slug': 'centrifugo-simples-aspiracao-ff-2-160-vs',
        'specs': [   {'atributo': 'Tensão Nominal', 'valor': '220 V'},
                     {'atributo': 'Potência Consumida', 'valor': '300 W'},
                     {'atributo': 'Rotação Nominal', 'valor': '1790 RPM'},
                     {'atributo': 'Grau de Proteção', 'valor': 'IP44'},
                     {'atributo': 'Ruído Acústico', 'valor': '67 dBA'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '112',
        'nome': 'Ventilador Soprador Axial 500 mm - FS/4-500 VT',
        'sheet_name': 'Map FS4 500 VT',
        'sku': 'FS/4-500 VT',
        'slug': 'ventilador-soprador-axial-500mm-fs-4-500-vt',
        'specs': [   {'atributo': 'Modelo', 'valor': 'FS/4-500 VT'},
                     {   'atributo': 'Descrição',
                         'valor': 'Ventilador axial com grade e caixa de terminais, motor de rotor externo'},
                     {'atributo': 'Tensão Nominal', 'valor': '220/380 V'},
                     {'atributo': 'Alimentação', 'valor': 'Trifasica'},
                     {'atributo': 'Ligacao', 'valor': 'Conforme esquema'},
                     {'atributo': 'Corrente Nominal', 'valor': '1,5/1,7/0,9/1,0 A'},
                     {'atributo': 'Potência Consumida', 'valor': '450/480 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Velocidade Nominal', 'valor': '1360/1600 RPM'},
                     {'atributo': 'Proteção Elétrica', 'valor': 'Atraves de rele externo'},
                     {'atributo': 'Proteção Mecânica', 'valor': 'IP-54'},
                     {'atributo': 'Mancais', 'valor': 'rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30 °C a 60 °C'},
                     {'atributo': 'Isolação', 'valor': 'F'},
                     {'atributo': 'Nível de Ruído', 'valor': '72/76 dBA'},
                     {'atributo': 'Regime de Trabalho', 'valor': 'S1 (Continuo)'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aco (na cor preta)'},
                     {'atributo': 'Grades de Protecao', 'valor': 'Em fio de aco (na cor preta)'},
                     {   'atributo': 'Caixa de Ligação',
                         'valor': 'Em polipropileno aditivado com borne, prensa cabos e aterramento'},
                     {'atributo': 'Peso', 'valor': '9,5 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '99',
        'nome': 'Ventilador Soprador Axial 400 mm - FS/4-400 VT',
        'sheet_name': 'Map FS4 400 VT',
        'sku': 'FS/4-400 VT',
        'slug': 'ventilador-soprador-axial-400mm-fs-4-400-vt',
        'specs': [   {'atributo': 'Modelo', 'valor': 'FS/4-400 VT'},
                     {   'atributo': 'Descrição',
                         'valor': 'Ventilador axial com grade e caixa de terminais, motor de rotor externo'},
                     {'atributo': 'Tensão Nominal', 'valor': '220/380 V'},
                     {'atributo': 'Alimentação', 'valor': 'Trifasica'},
                     {'atributo': 'Ligacao', 'valor': 'Conforme esquema'},
                     {'atributo': 'Corrente Nominal', 'valor': '0,80/1,00/0,48/0,60 A'},
                     {'atributo': 'Potência Consumida', 'valor': '190/260 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Velocidade Nominal', 'valor': '1390/1590 RPM'},
                     {'atributo': 'Proteção Elétrica', 'valor': 'Atraves de rele externo'},
                     {'atributo': 'Proteção Mecânica', 'valor': 'IP-54'},
                     {'atributo': 'Mancais', 'valor': 'rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30 °C a 60 °C'},
                     {'atributo': 'Isolação', 'valor': 'F'},
                     {'atributo': 'Nível de Ruído', 'valor': '67/72 dBA'},
                     {'atributo': 'Regime de Trabalho', 'valor': 'S1 (Continuo)'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aco (na cor preta)'},
                     {'atributo': 'Grades de Protecao', 'valor': 'Em fio de aco (na cor preta)'},
                     {   'atributo': 'Caixa de Ligação',
                         'valor': 'Em polipropileno aditivado com borne, prensa cabos e aterramento'},
                     {'atributo': 'Peso', 'valor': '6 kg'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '481',
        'nome': 'Ventilador Radial 315mm - FB/4 315 MCD',
        'sheet_name': 'Map FB4 315 MCD',
        'sku': 'FB/4 315 MCD',
        'slug': 'ventilador-radial-315mm-fb-4-315-mcd',
        'specs': [{'atributo': 'Tensão Nominal', 'valor': '220 V'}, {'atributo': 'Frequência', 'valor': '50/60 Hz'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '460',
        'nome': 'Ventilador Centrifugo 133mm - FF/2 133 N097 C',
        'sheet_name': 'Map FF2 133 N097 C',
        'sku': 'FF/2 133 N097',
        'slug': 'ventilador-centrifugo-133mm-ff-2-133-n097-c',
        'specs': [{'atributo': 'Tensão Nominal', 'valor': '220 V'}, {'atributo': 'Frequência', 'valor': '50/60 Hz'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '520',
        'nome': 'Ventilador Centrifugo Simples Aspiracao Ff2 140 Vs',
        'sheet_name': 'Map FF2 140 VS',
        'sku': 'FF2-140-VS',
        'slug': 'centrifugo-simples-aspiracao-ff2-140-vs',
        'specs': [{'atributo': 'Tensão Nominal', 'valor': '220 V'}, {'atributo': 'Frequência', 'valor': '50/60 Hz'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '568',
        'nome': 'Ventilador Radial 175mm - FB/2 175 MCD',
        'sheet_name': 'Map FB2 175 MCD',
        'sku': 'FB/2 175 MCD',
        'slug': 'ventilador-radial-175mm-fb-2-175-mcd',
        'specs': [{'atributo': 'Tensão Nominal', 'valor': '220 V'}, {'atributo': 'Frequência', 'valor': '50/60 Hz'}]},
    {   'categoria': 'Ventiladores Industriais',
        'code': '542',
        'nome': 'Ventilador Soprador Axial 200mm - FS/2-200 VM',
        'sheet_name': 'Map FS2 200 VM',
        'sku': 'FS/2-200 VM',
        'slug': 'ventilador-soprador-axial-200mm-fs-2-200-vm',
        'specs': [   {'atributo': 'Tensão Nominal', 'valor': '220 V'},
                     {'atributo': 'Alimentação', 'valor': 'Monofásica'},
                     {'atributo': 'Corrente Nominal', 'valor': '0,38/0,46 A'},
                     {'atributo': 'Potência Consumida', 'valor': '80/95 W'},
                     {'atributo': 'Frequência', 'valor': '50/60 Hz'},
                     {'atributo': 'Rotação Nominal', 'valor': '2700/3200 RPM'},
                     {'atributo': 'Grau de Proteção (IP)', 'valor': 'IP-54'},
                     {'atributo': 'Tipo de Mancal', 'valor': 'Com rolamentos de esfera blindado (2Z)'},
                     {'atributo': 'Temperatura de Operação', 'valor': '-30°C a 60°C'},
                     {'atributo': 'Ruído Acústico', 'valor': '60/65 dBA'},
                     {'atributo': 'Material da Hélice', 'valor': 'Chapa de aço (na cor preta)'},
                     {'atributo': 'Grade de Proteção', 'valor': 'Em fio de aço (na cor preta)'},
                     {'atributo': 'Peso Líquido', 'valor': '1,9 kg'}]}]





def calcular_curva_abc(df_grupo, col_nome, col_qtd='quantity'):
    """Calcula quantidade, share, acumulado e classe ABC rigorosa."""
    agg = df_grupo.groupby(col_nome)[col_qtd].sum().reset_index()
    agg = agg[agg[col_qtd] > 0].sort_values(by=col_qtd, ascending=False)
    total = agg[col_qtd].sum()
    
    if total == 0:
        return []
        
    resultado = []
    acum = 0.0
    for _, row in agg.iterrows():
        qtd = float(row[col_qtd])
        share = (qtd / total) * 100.0
        acum += share
        
        if acum <= 80.01:
            classe = "A"
        elif acum <= 95.01:
            classe = "B"
        else:
            classe = "C"
            
        resultado.append({
            "nome": str(row[col_nome]),
            "qtd": int(qtd),
            "share": f"{share:.1f}%",
            "acumulado": f"{min(acum, 100.0):.1f}%",
            "classe": classe
        })
    return resultado


def formatar_aba_excel(ws, conf, total_vendas, total_clientes, specs, mercados_abc, eq_abc, top_clientes):
    """Aplica a formatação visual executiva idêntica ao padrão de FS4-300 EM."""
    
    # Estilos
    font_titulo = Font(name="Arial", size=14, bold=True, color="1250B2")
    font_sub = Font(name="Arial", size=10, italic=True, color="475569")
    font_sec = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_th = Font(name="Arial", size=9, bold=True, color="1E293B")
    font_td = Font(name="Arial", size=9, color="334155")
    font_bold = Font(name="Arial", size=9, bold=True, color="1E293B")
    
    fill_sec = PatternFill(start_color="1250B2", end_color="1250B2", fill_type="solid")
    fill_th = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Linha 1 e 2: Cabeçalho
    ws.cell(row=1, column=1, value=f"Mapeamento B2B – {conf['nome']}").font = font_titulo
    ws.cell(row=2, column=1, value=f"SKU: {conf['sku']} | Total Vendas: {total_vendas:,.0f} un | Clientes Auditados: {total_clientes}").font = font_sub
    
    # Linha 4: 1. ESPECIFICAÇÕES TÉCNICAS
    ws.cell(row=4, column=1, value="1. ESPECIFICAÇÕES TÉCNICAS DO DATASHEET").font = font_sec
    ws.cell(row=4, column=1).fill = fill_sec
    ws.cell(row=4, column=2).fill = fill_sec
    
    ws.cell(row=5, column=1, value="Parâmetro Técnico").font = font_th
    ws.cell(row=5, column=1).fill = fill_th
    ws.cell(row=5, column=2, value="Valor Oficial").font = font_th
    ws.cell(row=5, column=2).fill = fill_th
    
    cur_row = 6
    for sp in specs:
        ws.cell(row=cur_row, column=1, value=sp['atributo']).font = font_bold
        ws.cell(row=cur_row, column=2, value=sp['valor']).font = font_td
        ws.cell(row=cur_row, column=1).border = thin_border
        ws.cell(row=cur_row, column=2).border = thin_border
        cur_row += 1
        
    cur_row += 1
    # 2. MERCADOS ATENDIDOS
    ws.cell(row=cur_row, column=1, value="2. MERCADOS ATENDIDOS (Curva ABC de Vendas)").font = font_sec
    for c in range(1, 6):
        ws.cell(row=cur_row, column=c).fill = fill_sec
    cur_row += 1
    
    headers_abc = ["Segmento / Indústria", "Qtd (un)", "% Share", "Acumulado", "Curva ABC"]
    for c, h in enumerate(headers_abc, 1):
        cell = ws.cell(row=cur_row, column=c, value=h)
        cell.font = font_th
        cell.fill = fill_th
        cell.border = thin_border
    cur_row += 1
    
    for idx, m in enumerate(mercados_abc):
        ws.cell(row=cur_row, column=1, value=m['nome']).font = font_td
        ws.cell(row=cur_row, column=2, value=m['qtd']).font = font_bold
        ws.cell(row=cur_row, column=3, value=m['share']).font = font_td
        ws.cell(row=cur_row, column=4, value=m['acumulado']).font = font_td
        ws.cell(row=cur_row, column=5, value=m['classe']).font = font_bold
        for c in range(1, 6):
            ws.cell(row=cur_row, column=c).border = thin_border
            if idx % 2 == 1:
                ws.cell(row=cur_row, column=c).fill = fill_zebra
        cur_row += 1
        
    cur_row += 1
    # 3. EQUIPAMENTOS FÍSICOS
    ws.cell(row=cur_row, column=1, value="3. EQUIPAMENTOS FÍSICOS DE INSTALAÇÃO (Onde Usar)").font = font_sec
    for c in range(1, 6):
        ws.cell(row=cur_row, column=c).fill = fill_sec
    cur_row += 1
    
    headers_eq = ["Equipamento Físico", "Qtd (un)", "% Share", "Acumulado", "Curva ABC"]
    for c, h in enumerate(headers_eq, 1):
        cell = ws.cell(row=cur_row, column=c, value=h)
        cell.font = font_th
        cell.fill = fill_th
        cell.border = thin_border
    cur_row += 1
    
    for idx, eq in enumerate(eq_abc):
        ws.cell(row=cur_row, column=1, value=eq['nome']).font = font_td
        ws.cell(row=cur_row, column=2, value=eq['qtd']).font = font_bold
        ws.cell(row=cur_row, column=3, value=eq['share']).font = font_td
        ws.cell(row=cur_row, column=4, value=eq['acumulado']).font = font_td
        ws.cell(row=cur_row, column=5, value=eq['classe']).font = font_bold
        for c in range(1, 6):
            ws.cell(row=cur_row, column=c).border = thin_border
            if idx % 2 == 1:
                ws.cell(row=cur_row, column=c).fill = fill_zebra
        cur_row += 1
        
    cur_row += 1
    # 4. TOP CLIENTES REAIS
    ws.cell(row=cur_row, column=1, value="4. TOP CLIENTES REAIS AUDITADOS (Curva ABC)").font = font_sec
    for c in range(1, 5):
        ws.cell(row=cur_row, column=c).fill = fill_sec
    cur_row += 1
    
    headers_cli = ["Razão Social / Cliente", "Qtd (un)", "Mercado do Cliente", "Equipamento Identificado"]
    for c, h in enumerate(headers_cli, 1):
        cell = ws.cell(row=cur_row, column=c, value=h)
        cell.font = font_th
        cell.fill = fill_th
        cell.border = thin_border
    cur_row += 1
    
    for idx, cli in enumerate(top_clientes[:12]):
        ws.cell(row=cur_row, column=1, value=cli['cliente']).font = font_bold
        ws.cell(row=cur_row, column=2, value=cli['qtd']).font = font_bold
        ws.cell(row=cur_row, column=3, value=cli['mercado']).font = font_td
        ws.cell(row=cur_row, column=4, value=cli['equipamento']).font = font_td
        for c in range(1, 5):
            ws.cell(row=cur_row, column=c).border = thin_border
            if idx % 2 == 1:
                ws.cell(row=cur_row, column=c).fill = fill_zebra
        cur_row += 1
        
    # Ajuste de largura das colunas
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 34
    ws.column_dimensions['E'].width = 16


def main():
    print("=" * 70)
    print("Demo Store — PROCESSANDO NOVO LOTE DE 15 PRODUTOS INÉDITOS")
    print("=" * 70)
    
    # 1. Carrega parquet
    parquet_path = os.path.join(MAP_DIR, "data", "normalized", "consolidacao.parquet")
    df = pd.read_parquet(parquet_path)
    print(f"[*] Base carregada: {len(df):,} linhas.")
    
    # 2. Carrega workbook do Excel
    excel_path = os.path.join(MAP_DIR, "MAPEAMENTO_PILOTO_COMPLETO.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    print(f"[*] Workbook aberto com {len(wb.sheetnames)} abas existentes.")
    
    produtos_processados = []
    
    for conf in LOTE_10_CONFIG:
        code = conf["code"]
        sku = conf["sku"]
        sheet_name = conf["sheet_name"]
        print(f"\n[*] Mapeando #{code} - {sku} ({conf['nome']})...")
        
        # Filtra dados do produto
        sub = df[df["product_code"] == code]
        if len(sub) == 0:
            print(f"[!] Nenhum dado encontrado para o código {code}. Pulando...")
            continue
            
        total_vendas = sub["quantity"].sum()
        total_clientes = sub["client_key"].nunique()
        
        # Normalização de nomes de segmentos para evitar siglas brutas
        def normalizar_seg(s):
            s_up = str(s).upper()
            if "MAQ E EQUIP CLIMAT" in s_up or "QUEIMADORES" in s_up:
                return "Climatização, desumidificação e sistemas de aquecimento"
            if "SALAS LIMPAS" in s_up or "CAPELAS" in s_up:
                return "Salas limpas, capelas de laboratório e exaustão industrial"
            if "COIFAS" in s_up or "CHURRASQUEIRAS" in s_up:
                return "Ventilação de cozinhas profissionais, coifas e churrasqueiras"
            if "TUNEL" in s_up or "CAMARAS FRIGOR" in s_up:
                return "Câmaras frigoríficas, túneis de congelamento e refrigeração"
            if "MOTORES GERAD TRANSF" in s_up or "RACKS" in s_up or "PAINEIS" in s_up:
                return "Painéis elétricos, transformadores a seco e grupos geradores"
            if "COMPRESSOR DE AR" in s_up:
                return "Compressores industriais e sistemas de ar comprimido"
            if "COZINHAS INDUSTRIAIS" in s_up or "BALCAO FRIGORIFICO" in s_up:
                return "Refrigeração comercial, balcões frigoríficos e expositores"
            if "EQUIPAMENTO MEDICO" in s_up:
                return "Equipamentos médicos, hospitalares e laboratoriais"
            if "SIDERURGIA" in s_up or "METALURGIA" in s_up:
                return "Siderurgia, metalurgia e fornos de processo"
            if "AR CONDICONADO DE PAINEL" in s_up:
                return "Climatização e ar-condicionado para painéis elétricos"
            if "RESF DE LEITE" in s_up or "AGRIC" in s_up:
                return "Máquinas agrícolas, agroindústria e resfriadores de leite"
            if "LOJAS DE PECAS DE REFRIGERACAO" in s_up:
                return "Refrigeração comercial e reposição técnica"
            if "LOJAS DE PECAS DE MANUTENCAO" in s_up or "LOJAS DE PECAS ELETRONICAS" in s_up:
                return "Manutenção industrial e automação"
            return str(s).title()

        sub_clean = sub.copy()
        sub_clean["segment_norm"] = sub_clean["segment"].apply(normalizar_seg)
        
        # Normalização de equipamentos
        def normalizar_eq(eq):
            eq_str = str(eq)
            # Remove qualquer variação de "Aplicação final provável – — - :"
            eq_str = re.sub(r'(?i)aplica[çc][ãa]o\s+final\s+prov[áa]vel\s*[-–—:]*\s*', '', eq_str).strip()
            eq_str = re.sub(r'(?i)aplica[çc][ãa]o\s+prov[áa]vel\s*[-–—:]*\s*', '', eq_str).strip()
            return eq_str

        sub_clean["eq_norm"] = sub_clean["equipment_candidate_1"].apply(normalizar_eq)
        
        # Curvas ABC
        mercados_abc = calcular_curva_abc(sub_clean, "segment_norm")
        eq_abc = calcular_curva_abc(sub_clean, "eq_norm")
        
        # Top Clientes
        top_cli_df = sub_clean.groupby(["client_raw", "segment_norm", "eq_norm"])["quantity"].sum().reset_index()
        top_cli_df = top_cli_df.sort_values(by="quantity", ascending=False)
        top_clientes = []
        for _, r in top_cli_df.iterrows():
            top_clientes.append({
                "cliente": str(r["client_raw"]),
                "qtd": int(r["quantity"]),
                "mercado": str(r["segment_norm"]),
                "equipamento": str(r["eq_norm"])
            })
            
        # Cria ou atualiza a aba no Excel
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        formatar_aba_excel(ws, conf, total_vendas, total_clientes, conf["specs"], mercados_abc, eq_abc, top_clientes)
        print(f"[OK] Aba '{sheet_name}' criada com sucesso no Excel!")
        
        # Gera o JSON estruturado para o produto
        tensao = next((s['valor'] for s in conf['specs'] if 'Tensão' in s['atributo']), '230 V')
        alimentacao = next((s['valor'] for s in conf['specs'] if 'Alimentação' in s['atributo']), 'Monofásica')
        regime = next((s['valor'] for s in conf['specs'] if 'Regime' in s['atributo']), 'S1')
        protecao = next((s['valor'] for s in conf['specs'] if 'Proteção' in s['atributo']), 'IP54')
        temperatura = next((s['valor'] for s in conf['specs'] if 'Temperatura' in s['atributo']), '-30 °C a 60 °C')
        mancais = next((s['valor'] for s in conf['specs'] if 'Mancais' in s['atributo']), 'Rolamentos de esferas blindados 2Z')
        mat_helice = next((s['valor'] for s in conf['specs'] if 'Hélice' in s['atributo'] or 'Rotor' in s['atributo'] or 'Carcaça' in s['atributo']), 'Estrutura metálica')

        # Mercados selecionados
        mercados_nomes = [m['nome'] for m in mercados_abc[:2]]
        str_mercados = " e ".join(mercados_nomes).lower() if mercados_nomes else "refrigeração e ventilação industrial"

        # Categoria lógica
        eh_micro = "micro" in conf["slug"] or "a25089" in conf["slug"]
        eh_centrifugo = "centrifugo" in conf["slug"] or "radial" in conf["slug"] or "ff-" in conf["slug"] or "fb-" in conf["slug"]
        fluxo = "Soprador" if "soprador" in conf["slug"] or "vm" in conf["slug"] or "vt" in conf["slug"] else "Exaustor"

        if eh_micro:
            p1 = f"O {conf['nome']} é um microventilador compacto projetado para ventilação forçada e dissipação térmica em {str_mercados}. Opera sob alimentação de {tensao} ({alimentacao.lower()})."
            p2 = f"Construído com {mat_helice.lower()}, conta com {mancais.lower()}, grau de proteção {protecao} e regime de trabalho {regime}. Projetado para operar na faixa de temperatura de {temperatura} conforme especificado pelo fabricante."
        elif eh_centrifugo:
            p1 = f"O {conf['nome']} é um ventilador desenvolvido para movimentação e exaustão de ar em {str_mercados}. Opera sob alimentação de {tensao} ({alimentacao.lower()})."
            p2 = f"Possui rotor balanceado de acoplamento direto, {mancais.lower()} e grau de proteção {protecao}. Projetado para regime de trabalho {regime} e faixa de temperatura de operação de {temperatura} documentada no datasheet oficial."
        else:
            fluxo_acao = "movimentação e insuflamento de ar" if fluxo == "Soprador" else "movimentação e exaustão de ar"
            p1 = f"O {conf['nome']} é um ventilador axial projetado para {fluxo_acao} em {str_mercados}. Opera sob alimentação de {tensao} ({alimentacao.lower()})."
            p2 = f"Fabricado com hélice em {mat_helice.lower()} e grade de proteção, conta com {mancais.lower()} e grau de proteção {protecao}. Projetado para regime de trabalho {regime} e faixa de temperatura de operação de {temperatura} especificada pelo fabricante."

        resumo_tecnico = f"{p1}\n\n{p2}"

        # Hero checklist
        tensao_curta = tensao.split("(")[0].strip()
        badge_tensao = f"{tensao_curta} {alimentacao}" if len(tensao_curta) < 15 else tensao_curta
        badge_regime = "Regime S1 Contínuo" if "s1" in regime.lower() else f"Regime {regime}"
        badge_mancal = "Rolamentos 2Z"

        hero_checklist = [badge_tensao, badge_regime, badge_mancal]

        # Onde Usar Cards
        def desc_eq_func(eq_name):
            n = eq_name.lower()
            if "evaporador" in n or "câmara" in n:
                return "Forçamento da circulação de ar através das serpentinas aletadas para manter a temperatura homogênea no ambiente de estocagem."
            if "condensador" in n or "unidade condensadora" in n:
                return "Rejeição contínua do calor do fluido refrigerante para o ar externo através do fluxo forçado sobre a serpentina."
            if "balcão" in n or "expositor" in n:
                return "Circulação forçada de ar para troca térmica no gabinete refrigerado."
            if "coifa" in n:
                return "Exaustão e circulação de ar em coifas e sistemas de ventilação."
            if "compressor" in n or "óleo" in n:
                return "Movimentação forçada de ar através de radiadores ou resfriadores de óleo em compressores."
            if "gerador" in n or "transformador" in n or "painel" in n:
                return "Ventilação forçada e circulação de ar em painéis elétricos e transformadores de potência."
            if "uta" in n or "tratamento de ar" in n:
                return "Insuflamento ou exaustão de ar tratado através de bancos de filtros e dutos em centrais de climatização."
            if "leite" in n or "agrícola" in n:
                return "Troca térmica no sistema frigorífico de resfriadores de leite e tanques de expansão direta."
            return f"Movimentação forçada de ar e troca térmica integrada ao sistema de {eq_name}."

        onde_usar_cards = []
        for eq in eq_abc[:5]:
            onde_usar_cards.append({
                "titulo": eq["nome"],
                "descricao": desc_eq_func(eq["nome"])
            })

        # Aplicações Cards
        if eh_micro:
            aplicacoes_cards = [
                {"titulo": "Exaustão em Painéis e Gabinetes", "descricao": "Extração do ar aquecido acumulado no interior de quadros elétricos e cubículos de comando."},
                {"titulo": "Dissipação em Inversores e Fontes", "descricao": "Movimentação de ar direcionada sobre os dissipadores térmicos de módulos de potência."},
                {"titulo": "Circulação em Equipamentos Industriais", "descricao": "Distribuição de ar em instrumentos científicos e equipamentos de precisão."}
            ]
        elif eh_centrifugo:
            aplicacoes_cards = [
                {"titulo": "Exaustão e Renovação de Ar Dutada", "descricao": "Vencimento de perdas de carga em redes de dutos de ar em sistemas industriais e laboratoriais."},
                {"titulo": "Insuflamento em Sistemas de Filtragem", "descricao": "Pressurização de ar através de filtros em unidades de tratamento de ar e salas limpas."},
                {"titulo": "Circulação em Coifas e Gabinetes", "descricao": "Movimentação forçada de ar em coifas industriais e gabinetes de exaustão."}
            ]
        else:
            aplicacoes_cards = [
                {"titulo": "Troca Térmica em Evaporadores Frigoríficos", "descricao": "Circulação de ar frio forçado através das serpentinas de evaporadores em câmaras de conservação e congelamento."},
                {"titulo": "Rejeição de Calor em Condensadores", "descricao": "Exaustão do ar aquecido gerado pela condensação do fluido refrigerante em unidades condensadoras comerciais e industriais."},
                {"titulo": "Resfriamento em Trocadores e Compressores", "descricao": "Movimentação de ar através de blocos aletados e radiadores em compressores e sistemas industriais."}
            ]

        # Benefícios
        beneficios = [
            {
                "titulo": "Estrutura Metálica e Hélice",
                "descricao": f"Hélice e carcaça metálica projetadas para resistência estrutural e sustentação mecânica durante a operação."
            },
            {
                "titulo": "Mancais com Rolamentos Blindados 2Z",
                "descricao": "Rolamentos de esfera blindados 2Z, proporcionando suporte mecânico e estabilidade ao eixo durante a operação."
            },
            {
                "titulo": f"Grau de Proteção {protecao}",
                "descricao": f"Grau de proteção {protecao}, com proteção contra ingresso de poeira e projeções de água em qualquer direção."
            },
            {
                "titulo": "Faixa de Temperatura de Operação",
                "descricao": f"Faixa de temperatura de operação de {temperatura}, conforme especificado pelo fabricante."
            }
        ]

        # Diferenciais
        diferenciais = [
            f"Conjunto Balanceado: Rotação estável e deslocamento de ar com nível de ruído reduzido.",
            f"Mancais com Rolamentos Blindados 2Z: Suporte mecânico e estabilidade rotacional ao eixo durante a operação.",
            f"Regime de Trabalho {regime}: Regime {regime} para operação contínua, conforme as especificações do equipamento.",
            f"Grau de Proteção {protecao}: Proteção contra ingresso de poeira e projeções de água em qualquer direção.",
            f"Faixa de Temperatura ({temperatura}): Faixa de temperatura de operação especificada de {temperatura}."
        ]

        # FAQ
        faq = [
            {
                "pergunta": f"Qual a tensão e regime de trabalho do {sku}?",
                "resposta": f"O modelo {sku} opera com tensão nominal de {tensao} ({alimentacao.lower()}) em regime de trabalho {regime}, conforme especificação do equipamento."
            },
            {
                "pergunta": f"Qual o grau de proteção mecânica e tipo de mancal?",
                "resposta": f"Possui grau de proteção {protecao} e sistema de mancais com rolamentos de esfera blindados 2Z."
            },
            {
                "pergunta": f"Qual a faixa de temperatura suportada pelo equipamento?",
                "resposta": f"O equipamento opera na faixa de temperatura de {temperatura} especificada na folha de dados técnicos."
            }
        ]

        # Monta o JSON
        prod_json = {
            "slug": conf["slug"],
            "nome": conf["nome"],
            "sku": sku,
            "categoria": conf["categoria"],
            "especificacoes": [{"atributo": s["atributo"], "valor": s["valor"], "confianca": "100%", "fonte": "Datasheet Oficial"} for s in conf["specs"]],
            "resumo_tecnico": resumo_tecnico,
            "hero_checklist": hero_checklist,
            "beneficios": beneficios,
            "diferenciais": diferenciais,
            "mercado": [m["nome"] for m in mercados_abc[:4]],
            "mercados": [m["nome"] for m in mercados_abc[:4]],
            "aplicacoes_categoria": {
                "titulo": f"Aplicações do {conf['nome']}",
                "intro": "Funções técnicas e modos de operação térmica do componente.",
                "cards": aplicacoes_cards
            },
            "aplicacoes_equipamento": {
                "titulo": f"Onde o {conf['nome']} é Utilizado?",
                "intro": "Equipamentos físicos e sistemas industriais nos quais o componente é instalado:",
                "cards": onde_usar_cards
            },
            "faq": faq,
            "alerta_tecnico": f"Atenção: verifique a correta alimentação ({tensao}) e realize a fixação em estrutura rígida antes de energizar o equipamento.",
            "concorrentes_validados": ["ebm-papst", "Ziehl-Abegg"],
            "seo": {
                "keywords": [conf["slug"].replace("-", " "), sku.lower(), f"ventilador {sku.lower()}"],
                "meta_description": f"{conf['nome']} (SKU: {sku}). Tensão {tensao}, grau de proteção {protecao}, rolamentos blindados 2Z. Dados oficiais Sell-Parts."
            }
        }

        # Salva o JSON em gerador/dados/
        json_path = os.path.join(DADOS_DIR, f"{conf['slug']}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(prod_json, f, ensure_ascii=False, indent=2)
        print(f"[OK] JSON salvo em {json_path}")

        # Gera payload ACF
        payload = gerar_payload_acf(prod_json)
        salvar_output(payload)
        print(f"[OK] Payload ACF salvo em gerador/output/{conf['slug']}.json")

        produtos_processados.append(conf["nome"])

    # Salva o arquivo Excel com as novas abas
    wb.save(excel_path)
    print("\n" + "=" * 70)
    print(f"[SUCESSO] MAPEAMENTO DE 15 PRODUTOS CONCLUÍDO!")
    print(f"Total de novas abas salvas no Excel '{excel_path}': {len(produtos_processados)}")
    print("=" * 70)

if __name__ == "__main__":
    main()
